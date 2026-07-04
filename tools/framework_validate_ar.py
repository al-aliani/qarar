import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl


REQUIRED_SHEETS_AR = [
    "الافتراضات",
    "الاستثمار الأولي",
    "تقدير الإيرادات",
    "تقدير التكاليف",
    "قائمة الدخل",
    "التدفقات النقدية",
    "مؤشرات التقييم",
    "تحليل التعادل",
    "فترة الاسترداد",
    "تحليل الحساسية",
    "السيناريوهات",
    "مصفوفة المخاطر",
    "الملخص التنفيذي",
    "لوحة التحكم",
]

ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NUM!")
VOLATILE_FUNCS = ("NOW(", "TODAY(", "RAND(", "RANDBETWEEN(", "OFFSET(", "INDIRECT(", "CELL(", "INFO(")


@dataclass
class Finding:
    level: str  # "ERROR" | "WARN"
    code: str
    message: str
    details: Dict[str, Any]


def _safe_str(v: Any) -> str:
    try:
        return "" if v is None else str(v)
    except Exception:
        return "<unprintable>"


def _trim_map(sheetnames: List[str]) -> Dict[str, List[str]]:
    m: Dict[str, List[str]] = {}
    for s in sheetnames:
        m.setdefault(s.strip(), []).append(s)
    return m


def _match_required_sheets(sheetnames: List[str]) -> Tuple[List[str], Dict[str, str]]:
    """
    Returns: (missing_required, resolved_map)
    resolved_map maps required canonical name -> actual sheet name found (trim-insensitive).
    """
    trim_to_actual = _trim_map(sheetnames)
    resolved: Dict[str, str] = {}
    missing: List[str] = []
    for req in REQUIRED_SHEETS_AR:
        candidates = trim_to_actual.get(req.strip(), [])
        if not candidates:
            missing.append(req)
        else:
            # if multiple, pick first but still validator will flag collision elsewhere
            resolved[req] = candidates[0]
    return missing, resolved


def validate(path: Path) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(str(path), data_only=False)
    findings: List[Finding] = []

    # Sheet name hygiene (Trim)
    trim_map = _trim_map(wb.sheetnames)
    for s in wb.sheetnames:
        if s != s.strip():
            findings.append(
                Finding(
                    level="WARN",
                    code="SHEET_NAME_HAS_TRIM",
                    message="Sheet name has leading/trailing spaces (can break references).",
                    details={"sheet": s, "trimmed": s.strip()},
                )
            )
    for key, variants in trim_map.items():
        if len(variants) > 1:
            findings.append(
                Finding(
                    level="WARN",
                    code="SHEET_NAME_TRIM_COLLISION",
                    message="Duplicate sheet names after trim (high risk).",
                    details={"trimmed": key, "variants": variants},
                )
            )

    missing, resolved = _match_required_sheets(wb.sheetnames)
    if missing:
        findings.append(
            Finding(
                level="ERROR",
                code="MISSING_REQUIRED_SHEETS_AR",
                message="Missing required Arabic framework sheets.",
                details={"missing": missing, "expected_count": len(REQUIRED_SHEETS_AR)},
            )
        )

    external_refs = []
    broken_ref_in_formula = []
    volatile_cells = []
    error_values = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        dim = ws.calculate_dimension()
        for row in ws[dim]:
            for cell in row:
                v = cell.value
                if isinstance(v, str) and any(v.startswith(tok) for tok in ERROR_TOKENS):
                    error_values.append({"sheet": sname, "cell": cell.coordinate, "value": v})

                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                up = v.upper()
                if "[" in up and "]" in up:
                    external_refs.append({"sheet": sname, "cell": cell.coordinate, "formula": v})
                if "#REF!" in up:
                    broken_ref_in_formula.append({"sheet": sname, "cell": cell.coordinate, "formula": v})
                if any(fn in up for fn in VOLATILE_FUNCS):
                    volatile_cells.append({"sheet": sname, "cell": cell.coordinate, "formula": v})

    if external_refs:
        findings.append(
            Finding(
                level="ERROR",
                code="EXTERNAL_REFERENCES",
                message="External workbook references found in formulas.",
                details={"count": len(external_refs), "sample": external_refs[:30]},
            )
        )
    if broken_ref_in_formula:
        findings.append(
            Finding(
                level="ERROR",
                code="BROKEN_REF_IN_FORMULA",
                message="Found #REF! inside formulas (broken internal references).",
                details={"count": len(broken_ref_in_formula), "sample": broken_ref_in_formula[:30]},
            )
        )
    if error_values:
        findings.append(
            Finding(
                level="ERROR",
                code="ERROR_VALUES",
                message="Found Excel error values in cells.",
                details={"count": len(error_values), "sample": error_values[:30]},
            )
        )
    if volatile_cells:
        findings.append(
            Finding(
                level="WARN",
                code="VOLATILE_FUNCTIONS",
                message="Volatile functions detected.",
                details={"count": len(volatile_cells), "sample": volatile_cells[:20]},
            )
        )

    verdict = "FAIL" if any(f.level == "ERROR" for f in findings) else "PASS"
    return {
        "file": str(path),
        "sheet_count": len(wb.sheetnames),
        "sheets": wb.sheetnames,
        "resolved_required_sheets": resolved,
        "verdict": verdict,
        "findings": [f.__dict__ for f in findings],
    }


def main(argv: List[str]) -> int:
    if len(argv) < 2:
        print("Usage: python tools/framework_validate_ar.py <file.xlsx|folder>")
        return 2

    target = Path(argv[1]).expanduser()
    results: List[Dict[str, Any]] = []

    if target.is_dir():
        for p in sorted(target.glob("*.xlsx")):
            results.append(validate(p))
    else:
        results.append(validate(target))

    print(json.dumps({"results": results}, ensure_ascii=False, indent=2))
    return 3 if any(r["verdict"] != "PASS" for r in results) else 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))

