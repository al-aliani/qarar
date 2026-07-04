import sys
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl


REQUIRED_SHEETS = [
    "00_Cover",
    "01_Assumptions",
    "03_RevenueDrivers",
    "04_Opex",
    "05_Capex",
    "10_IncomeStatement",
    "11_CashFlow",
    "20_KPIs",
    "21_Scenarios",
    "22_Sensitivity",
    "30_Risks",
    "40_Dashboard",
    "99_Audit",
]

REQUIRED_NAMED_RANGES = [
    "assump_inflation",
    "assump_tax_rate",
    "assump_discount_rate",
    "assump_years",
    "assump_scenario",
]

VOLATILE_FUNCS = ("NOW(", "TODAY(", "RAND(", "RANDBETWEEN(", "OFFSET(", "INDIRECT(", "CELL(", "INFO(")
ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NUM!")


@dataclass
class Finding:
    level: str  # "ERROR" | "WARN"
    code: str
    message: str
    details: Dict[str, Any]


def _safe_str(v: Any) -> str:
    try:
        return "" if v is None else str(v)
    except Exception:
        return "<unprintable>"


def _iter_named_ranges(wb) -> Dict[str, str]:
    out: Dict[str, str] = {}
    try:
        for dn in wb.defined_names.definedName:
            out[dn.name] = dn.attr_text or ""
    except Exception:
        # Best-effort only
        pass
    return out


def _named_range_single_cell_value(wb, name: str) -> Tuple[bool, Any, str]:
    """
    Tries to resolve a defined name to a single cell value.
    Returns (ok, value, ref)
    """
    dn = wb.defined_names.get(name)
    if dn is None:
        return False, None, ""
    destinations = list(dn.destinations)  # (sheet, "A1") etc.
    if len(destinations) != 1:
        return False, None, _safe_str(getattr(dn, "attr_text", ""))
    sheet, coord = destinations[0]
    if sheet not in wb.sheetnames:
        return False, None, f"{sheet}!{coord}"
    ws = wb[sheet]
    return True, ws[coord].value, f"{sheet}!{coord}"


def validate_workbook(path: Path) -> Dict[str, Any]:
    findings: List[Finding] = []

    wb = openpyxl.load_workbook(str(path), data_only=False)

    # 1) Sheet name hygiene
    trimmed = {}
    for s in wb.sheetnames:
        key = s.strip()
        trimmed.setdefault(key, []).append(s)
        if s != key:
            findings.append(
                Finding(
                    level="ERROR",
                    code="SHEET_NAME_HAS_TRIM",
                    message="Sheet name has leading/trailing spaces (will break references).",
                    details={"sheet": s, "trimmed": key},
                )
            )
    for key, variants in trimmed.items():
        if len(variants) > 1:
            findings.append(
                Finding(
                    level="ERROR",
                    code="SHEET_NAME_TRIM_COLLISION",
                    message="Duplicate sheet names after trim.",
                    details={"trimmed": key, "variants": variants},
                )
            )

    # 2) Required sheets
    missing_sheets = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing_sheets:
        findings.append(
            Finding(
                level="ERROR",
                code="MISSING_REQUIRED_SHEETS",
                message="Workbook is missing required template sheets.",
                details={"missing": missing_sheets},
            )
        )

    # 3) Named ranges
    defined = _iter_named_ranges(wb)
    missing_names = [n for n in REQUIRED_NAMED_RANGES if n not in defined]
    if missing_names:
        findings.append(
            Finding(
                level="ERROR",
                code="MISSING_NAMED_RANGES",
                message="Required named ranges are missing.",
                details={"missing": missing_names},
            )
        )
    else:
        # Basic type checks for key assumptions
        for n in REQUIRED_NAMED_RANGES:
            ok, value, ref = _named_range_single_cell_value(wb, n)
            if not ok:
                findings.append(
                    Finding(
                        level="ERROR",
                        code="NAMED_RANGE_NOT_SINGLE_CELL",
                        message="Named range must resolve to a single cell.",
                        details={"name": n, "ref": ref},
                    )
                )
                continue
            if n in ("assump_inflation", "assump_tax_rate", "assump_discount_rate"):
                if not isinstance(value, (int, float)):
                    findings.append(
                        Finding(
                            level="ERROR",
                            code="ASSUMPTION_NOT_NUMERIC",
                            message="Key rate assumption must be numeric (not text).",
                            details={"name": n, "ref": ref, "value": _safe_str(value)},
                        )
                    )
            if n == "assump_years":
                if not isinstance(value, (int, float)) or int(value) <= 0:
                    findings.append(
                        Finding(
                            level="ERROR",
                            code="ASSUMPTION_INVALID_YEARS",
                            message="assump_years must be a positive number.",
                            details={"name": n, "ref": ref, "value": _safe_str(value)},
                        )
                    )

    # 4) Scan formulas and values
    external_refs: List[Dict[str, str]] = []
    volatile_cells: List[Dict[str, str]] = []
    broken_ref_in_formula: List[Dict[str, str]] = []
    error_values: List[Dict[str, str]] = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        dim = ws.calculate_dimension()
        # Iterate only used range by openpyxl's dimension
        for row in ws[dim]:
            for cell in row:
                v = cell.value

                # Error values in cells (including placeholder-like '#REF!' etc.)
                if isinstance(v, str) and any(v.startswith(tok) for tok in ERROR_TOKENS):
                    error_values.append({"sheet": sname, "cell": cell.coordinate, "value": v})

                if not (isinstance(v, str) and v.startswith("=")):
                    continue

                up = v.upper()
                if "[" in up and "]" in up:
                    external_refs.append({"sheet": sname, "cell": cell.coordinate, "formula": v})

                if any(fn in up for fn in VOLATILE_FUNCS):
                    volatile_cells.append({"sheet": sname, "cell": cell.coordinate, "formula": v})

                if "#REF!" in up:
                    broken_ref_in_formula.append({"sheet": sname, "cell": cell.coordinate, "formula": v})

    if external_refs:
        findings.append(
            Finding(
                level="ERROR",
                code="EXTERNAL_REFERENCES",
                message="External workbook references found in formulas.",
                details={"count": len(external_refs), "sample": external_refs[:20]},
            )
        )
    if broken_ref_in_formula:
        findings.append(
            Finding(
                level="ERROR",
                code="BROKEN_REF_IN_FORMULA",
                message="Found #REF! inside formulas (broken internal references).",
                details={"count": len(broken_ref_in_formula), "sample": broken_ref_in_formula[:30]},
            )
        )
    if error_values:
        findings.append(
            Finding(
                level="ERROR",
                code="ERROR_VALUES",
                message="Found Excel error values in cells.",
                details={"count": len(error_values), "sample": error_values[:30]},
            )
        )
    if volatile_cells:
        findings.append(
            Finding(
                level="WARN",
                code="VOLATILE_FUNCTIONS",
                message="Volatile functions detected; can break auditability and recalculation stability.",
                details={"count": len(volatile_cells), "sample": volatile_cells[:20]},
            )
        )

    # Verdict
    has_errors = any(f.level == "ERROR" for f in findings)
    return {
        "file": str(path),
        "sheet_count": len(wb.sheetnames),
        "verdict": "FAIL" if has_errors else "PASS",
        "findings": [f.__dict__ for f in findings],
    }


def main(argv: List[str]) -> int:
    if len(argv) < 2:
        print("Usage: python tools/feasibility_validate.py <file.xlsx|folder>")
        return 2

    target = Path(argv[1]).expanduser()
    results: List[Dict[str, Any]] = []

    if target.is_dir():
        for p in sorted(target.glob("*.xlsx")):
            results.append(validate_workbook(p))
    else:
        results.append(validate_workbook(target))

    # Print JSON (machine-friendly)
    print(json.dumps({"results": results}, ensure_ascii=False, indent=2))

    # Exit non-zero if any fail
    if any(r["verdict"] != "PASS" for r in results):
        return 3
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))

