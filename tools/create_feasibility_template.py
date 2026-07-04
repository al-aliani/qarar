import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


SHEETS = [
    "00_Cover",
    "01_Assumptions",
    "02_UnitEconomics",
    "03_RevenueDrivers",
    "04_Opex",
    "05_Capex",
    "06_WorkingCapital",
    "07_Financing",
    "10_IncomeStatement",
    "11_CashFlow",
    "12_BalanceSheet",
    "20_KPIs",
    "21_Scenarios",
    "22_Sensitivity",
    "30_Risks",
    "40_Dashboard",
    "99_Audit",
]


def _title(ws, text: str):
    ws["A1"] = text
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def _add_named(wb, name: str, sheet: str, cell: str):
    dn = DefinedName(name=name, attr_text=f"'{sheet}'!${cell}$")
    wb.defined_names.append(dn)


def build_template(out_path: Path):
    wb = openpyxl.Workbook()
    # Remove default
    wb.remove(wb.active)

    for s in SHEETS:
        wb.create_sheet(s)

    # Cover
    ws = wb["00_Cover"]
    _title(ws, "غلاف دراسة الجدوى")
    ws["A3"] = "اسم المشروع"
    ws["B3"] = ""
    ws["A4"] = "الإصدار"
    ws["B4"] = "0.1"
    ws["A5"] = "تاريخ التحديث"
    ws["B5"] = ""
    ws["A6"] = "صاحب الدراسة"
    ws["B6"] = ""
    ws["A8"] = "Scope (ما يغطيه النموذج)"
    ws["A9"] = "Out of Scope (ما لا يغطيه النموذج)"
    for r in range(3, 10):
        ws[f"A{r}"].font = Font(bold=True)

    # Assumptions
    ws = wb["01_Assumptions"]
    _title(ws, "الافتراضات الأساسية (Single Source of Truth)")
    ws["A3"] = "معدل التضخم"
    ws["B3"] = 0.02
    ws["A4"] = "معدل الضريبة"
    ws["B4"] = 0.15
    ws["A5"] = "معدل الخصم (Discount Rate)"
    ws["B5"] = 0.12
    ws["A6"] = "عدد سنوات النموذج"
    ws["B6"] = 5
    ws["A7"] = "السيناريو"
    ws["B7"] = "Base"

    hdr_fill = PatternFill("solid", fgColor="FF1F4E79")
    for c in ("A3", "A4", "A5", "A6", "A7"):
        ws[c].font = Font(bold=True, color="FFFFFFFF")
        ws[c].fill = hdr_fill
        ws[c].alignment = Alignment(horizontal="right")
    for c in ("B3", "B4", "B5", "B6", "B7"):
        ws[c].alignment = Alignment(horizontal="center")
    ws["B3"].number_format = "0%"
    ws["B4"].number_format = "0%"
    ws["B5"].number_format = "0%"

    dv = DataValidation(type="list", formula1='"Base,Best,Worst"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["B7"])

    # Named ranges (required by validator/spec)
    wb.defined_names.append(DefinedName("assump_inflation", attr_text="'01_Assumptions'!$B$3"))
    wb.defined_names.append(DefinedName("assump_tax_rate", attr_text="'01_Assumptions'!$B$4"))
    wb.defined_names.append(DefinedName("assump_discount_rate", attr_text="'01_Assumptions'!$B$5"))
    wb.defined_names.append(DefinedName("assump_years", attr_text="'01_Assumptions'!$B$6"))
    wb.defined_names.append(DefinedName("assump_scenario", attr_text="'01_Assumptions'!$B$7"))

    # KPI skeleton
    ws = wb["20_KPIs"]
    _title(ws, "المؤشرات وقرار الاستثمار")
    ws["A3"] = "NPV"
    ws["A4"] = "IRR"
    ws["A5"] = "Payback"
    ws["A6"] = "Breakeven"
    ws["A7"] = "Decision (GO/NO-GO/REVISE)"
    for r in range(3, 8):
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"] = ""

    # Audit sheet skeleton
    ws = wb["99_Audit"]
    _title(ws, "التدقيق (QA Gate)")
    ws["A3"] = "QA Checklist"
    ws["A5"] = "أشغّل المدقق: python tools/feasibility_validate.py <file.xlsx>"
    ws["A7"] = "ممنوع وجود: #REF! / مراجع خارجية / مدخلات نصية مكان أرقام"
    ws["A3"].font = Font(bold=True)

    # Set some column widths
    for s in SHEETS:
        ws = wb[s]
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 24

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main(argv) -> int:
    out = Path(argv[1]) if len(argv) > 1 else Path("templates/Feasibility_Template.xlsx")
    build_template(out)
    print(f"Created: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))

