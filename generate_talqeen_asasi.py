import datetime as dt
import re
from collections import Counter, defaultdict

import openpyxl
from openpyxl.utils.cell import range_boundaries


def is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def is_text(v) -> bool:
    return isinstance(v, str) and v.strip() != "" and not is_formula(v)


def sheet_display_name(s: str) -> str:
    # Preserve exact name, but also show trimmed to highlight whitespace-only differences.
    trimmed = s.strip()
    if trimmed != s:
        return f"{s} (ملاحظة: الاسم يحتوي فراغات زائدة؛ بعد trim يصبح: `{trimmed}`)"
    return s


def detect_title_cells(ws, max_row=8, max_col=20):
    # Heuristic: biggest merged cell within the top area that contains text.
    merged = list(ws.merged_cells.ranges)
    best = None
    for r in merged:
        if r.min_row <= max_row and r.min_col <= max_col:
            area = (r.max_row - r.min_row + 1) * (r.max_col - r.min_col + 1)
            v = ws.cell(r.min_row, r.min_col).value
            if is_text(v):
                if best is None or area > best["area"]:
                    best = {
                        "range": str(r),
                        "cell": ws.cell(r.min_row, r.min_col).coordinate,
                        "value": str(v).strip(),
                        "area": area,
                    }
    return best


def top_left_text_samples(ws, rows=35, cols=35, limit=30):
    items = []
    seen = set()
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            v = ws.cell(r, c).value
            if not is_text(v):
                continue
            txt = str(v).strip()
            # Skip ultra-short tokens
            if len(txt) <= 1:
                continue
            # Keep unique
            if txt in seen:
                continue
            seen.add(txt)
            if len(txt) > 160:
                txt = txt[:160] + "…"
            items.append((ws.cell(r, c).coordinate, txt))
            if len(items) >= limit:
                return items
    return items


SHEET_REF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_]+))!")


def extract_sheet_refs(formula: str):
    refs = []
    for m in SHEET_REF_RE.finditer(formula):
        name = m.group(1) or m.group(2)
        if name:
            refs.append(name)
    return refs


def analyze_workbook(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    # Duplicated sheet names after trim detection
    trim_map = defaultdict(list)
    for s in wb.sheetnames:
        trim_map[s.strip()].append(s)
    trimmed_collisions = {k: v for k, v in trim_map.items() if len(v) > 1}

    overall = {
        "sheet_count": len(wb.sheetnames),
        "sheetnames": wb.sheetnames,
        "trimmed_collisions": trimmed_collisions,
    }

    per_sheet = []
    dependency_edges = Counter()  # (src, dst) -> count
    external_refs = []
    errors_hash_value = []  # cells containing literal '#'

    for sname in wb.sheetnames:
        ws = wb[sname]
        dim = ws.calculate_dimension()
        min_col, min_row, max_col, max_row = range_boundaries(dim)

        formula_cells = 0
        cross_sheet_refs = Counter()

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v == "#":
                    errors_hash_value.append((sname, cell.coordinate))

                if not is_formula(v):
                    continue
                formula_cells += 1
                up = v.upper()
                if "[" in up and "]" in up:
                    external_refs.append((sname, cell.coordinate, v))

                for ref in extract_sheet_refs(v):
                    if ref != sname:
                        cross_sheet_refs[ref] += 1
                        dependency_edges[(sname, ref)] += 1

        title = detect_title_cells(ws)
        samples = top_left_text_samples(ws)

        # Navigation links: cells in top-left area containing common navigation text.
        nav_hits = []
        for r in range(1, 15):
            for c in range(1, 35):
                v = ws.cell(r, c).value
                if not is_text(v):
                    continue
                txt = str(v).strip()
                if ("الرجوع" in txt) or ("جدول المحتويات" in txt) or ("المحتويات" in txt):
                    nav_hits.append((ws.cell(r, c).coordinate, txt))

        per_sheet.append(
            {
                "name": sname,
                "dimension": dim,
                "rows": max_row - min_row + 1,
                "cols": max_col - min_col + 1,
                "merged_ranges": len(ws.merged_cells.ranges),
                "formula_cells": formula_cells,
                "cross_sheet_refs": dict(cross_sheet_refs.most_common(25)),
                "title_cell": title,
                "top_left_text_samples": samples,
                "nav_hits": nav_hits,
            }
        )

    overall["total_formula_cells"] = sum(s["formula_cells"] for s in per_sheet)
    overall["external_ref_cells"] = external_refs
    overall["hash_cells"] = errors_hash_value

    # Build a simple dependency graph summary
    dep_by_src = defaultdict(list)
    for (src, dst), cnt in dependency_edges.items():
        dep_by_src[src].append((dst, cnt))
    dep_by_src = {k: sorted(v, key=lambda x: -x[1]) for k, v in dep_by_src.items()}

    overall["dependencies"] = dep_by_src

    return {"overall": overall, "sheets": per_sheet}


def write_baseline_prompt_md(analysis: dict, out_path: str, xlsx_path: str):
    overall = analysis["overall"]
    sheets = analysis["sheets"]

    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def md(s: str) -> str:
        return s.replace("\r\n", "\n").replace("\r", "\n")

    lines = []
    lines.append(f"## التلقين الأساسي لإعادة بناء ملف دراسة الجدوى (Blueprint)\n\n")
    lines.append(f"- **وقت التوليد**: {now}\n")
    lines.append(f"- **ملف المصدر**: `{xlsx_path}`\n")
    lines.append(f"- **عدد الشيتات**: {overall['sheet_count']}\n")
    lines.append(f"- **إجمالي خلايا الصيغ**: {overall['total_formula_cells']}\n\n")

    lines.append("### الفكرة العامة (Architecture)\n")
    lines.append(
        "- الملف مبني كنظام **مدخلات → جداول وسيطة → مؤشرات نهائية** موزعة على شيتات متخصصة.\n"
        "- توجد شيتات لحسابات مالية (قائمة دخل/تعادل/مؤشرات) وشيتات وصفية (فنية/قانونية/تسويقية/موارد) وشيت/شيتين للتسعير.\n"
        "- الروابط بين الشيتات تتم عبر **صيغ Excel تربط خلايا محددة** (بدون مراجع خارجية لملفات أخرى بحسب الفحص).\n\n"
    )

    if overall.get("trimmed_collisions"):
        lines.append("### ملاحظة مهمّة: أسماء شيتات متكررة بسبب الفراغات\n")
        lines.append(
            "يوجد شيتات لها نفس الاسم بعد إزالة الفراغات من البداية/النهاية (Trim). "
            "عند إعادة البناء يجب **مطابقة الاسم حرفيًا** (بما في ذلك الفراغات) لأن الصيغ تعتمد على الاسم الدقيق.\n\n"
        )
        for k, variants in overall["trimmed_collisions"].items():
            lines.append(f"- بعد Trim: `{k}` ← الأسماء الفعلية: {', '.join(f'`{v}`' for v in variants)}\n")
        lines.append("\n")

    if overall.get("hash_cells"):
        lines.append("### ملاحظة جودة بيانات: خلايا قيمتها حرف `#`\n")
        lines.append(
            "تم رصد خلايا تحتوي قيمة حرف واحد `#` (ليست بالضرورة أخطاء Excel القياسية مثل `#DIV/0!`). "
            "غالبًا هي **علامة Placeholder/فاصل** في القالب.\n\n"
        )
        # Show a small sample
        for sname, coord in overall["hash_cells"][:25]:
            lines.append(f"- `{sname}`!`{coord}`\n")
        if len(overall["hash_cells"]) > 25:
            lines.append(f"- ... (إجمالي {len(overall['hash_cells'])} خلية)\n")
        lines.append("\n")

    lines.append("### ترتيب الشيتات (يُفضّل مطابقته)\n")
    for i, s in enumerate(overall["sheetnames"], 1):
        lines.append(f"- {i}. `{s}`\n")
    lines.append("\n")

    lines.append("### قواعد التنفيذ عند إعادة البناء (Implementation rules)\n")
    lines.append(
        "- **لا تغيّر أسماء الشيتات** ولا تضف/تحذف فراغات؛ أي تغيير سيكسر الصيغ التي تشير للشيت.\n"
        "- حافظ على **نطاقات الدمج (Merged Cells)** لأنها تُستخدم كعناوين وأزرار تنقّل داخل القالب.\n"
        "- حافظ على **أماكن الجداول داخل نفس الأبعاد تقريبًا** (Dimension) لأن الصيغ غالبًا مبنية على خلايا ثابتة.\n"
        "- إدخال البيانات يكون عادةً في خلايا **بدون صيغة** داخل جداول، بينما الخلايا ذات الصيغ تمثل مخرجات/حسابات.\n\n"
    )

    lines.append("### خريطة الترابط بين الشيتات (Dependencies)\n")
    if overall.get("dependencies"):
        lines.append(
            "يوضح التالي أكثر الشيتات التي تُستَخدم كمصادر في صيغ شيتات أخرى (العدد = مرات الإشارة داخل الصيغ):\n\n"
        )
        for src, deps in overall["dependencies"].items():
            top = deps[:6]
            if not top:
                continue
            lines.append(f"- `{src}` يعتمد على: " + ", ".join(f"`{dst}`({cnt})" for dst, cnt in top) + "\n")
        lines.append("\n")
    else:
        lines.append("- لم يتم رصد ترابطات بين الشيتات داخل الصيغ.\n\n")

    lines.append("### مواصفات كل شيت (Sheet-by-sheet specification)\n")
    for idx, s in enumerate(sheets, 1):
        sname = s["name"]
        lines.append(f"\n## {idx}) {sheet_display_name(sname)}\n")
        lines.append(f"- **Dimension (نطاق الاستخدام)**: `{s['dimension']}` (تقريبًا {s['rows']} صف × {s['cols']} عمود)\n")
        lines.append(f"- **Merged ranges**: {s['merged_ranges']}\n")
        lines.append(f"- **خلايا صيغ**: {s['formula_cells']}\n")

        if s.get("title_cell"):
            t = s["title_cell"]
            lines.append(f"- **عنوان/هيدر مرجّح**: `{t['cell']}` ضمن `{t['range']}` = **{t['value']}**\n")

        if s.get("nav_hits"):
            lines.append("- **تنقّل داخل القالب** (روابط/أزرار نصية ظاهرة):\n")
            for coord, txt in s["nav_hits"][:10]:
                lines.append(f"  - `{coord}`: {txt}\n")

        if s.get("cross_sheet_refs"):
            lines.append("- **مصادر الصيغ من شيتات أخرى (Top)**:\n")
            for ref, cnt in list(s["cross_sheet_refs"].items())[:10]:
                lines.append(f"  - `{ref}`: {cnt}\n")

        # Provide top-left samples as “labels / headers clues”
        samples = s.get("top_left_text_samples") or []
        if samples:
            lines.append("- **نصوص/عناوين في أعلى الشيت (لقراءة شكل الجدول)**:\n")
            for coord, txt in samples[:18]:
                lines.append(f"  - `{coord}`: {txt}\n")

        # Implementation steps template
        lines.append("- **طريقة بناء شيت مشابه (قالب تنفيذ)**:\n")
        lines.append(
            "  - أنشئ العنوان في أعلى الشيت (غالبًا داخل خلية/نطاق مدمج) مع تنسيق خط أكبر/غامق.\n"
            "  - أنشئ صف/صفوف عناوين الأعمدة (Headers) ثم منطقة بيانات (Data area).\n"
            "  - ضع خلايا الإدخال كقيم مباشرة (بدون `=`) وخلايا النتائج بصيغ (`=`) تربط بباقي الشيتات.\n"
            "  - ثبّت نفس مواقع الخلايا المرجعية (Anchors) التي تُشار إليها من شيتات أخرى.\n"
        )

    lines.append("\n### مخرجات هذا التوليد\n")
    lines.append(
        "- هذا الملف هو **وصف تقني/Blueprint** لإعادة بناء القالب، وليس بديلًا عن البيانات.\n"
        "- للاطلاع على **كل خلية** (قيمة/صيغة/تنسيق) استخدم ملف التفريغ: `cells_dump_mac_blash.csv`.\n"
    )

    with open(out_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(md("".join(lines)))


if __name__ == "__main__":
    raise SystemExit("Run via runner (paths provided by assistant).")

