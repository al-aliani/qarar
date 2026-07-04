#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
إنشاء ملف Excel يحتوي على جميع جداول المنصة
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# تعريفات الجداول من schema.js
TABLES = {
    'buildings': {
        'title': 'المباني والإنشاءات',
        'category': 'CAPEX - الدراسة الفنية',
        'columns': ['البند', 'العدد', 'السعر', 'الإجمالي', '% استهلاك', 'ملاحظات']
    },
    'equipment': {
        'title': 'المعدات والأجهزة',
        'category': 'CAPEX - الدراسة الفنية',
        'columns': ['البند', 'العدد', 'السعر', 'الإجمالي', '% استهلاك', 'ملاحظات']
    },
    'furniture': {
        'title': 'الأثاث والتجهيزات',
        'category': 'CAPEX - الدراسة الفنية',
        'columns': ['البند', 'العدد', 'السعر', 'الإجمالي', '% استهلاك', 'ملاحظات']
    },
    'techResources': {
        'title': 'الموارد التقنية',
        'category': 'CAPEX - الموارد التقنية',
        'columns': ['البند', 'العدد', 'القيمة', 'الإجمالي', 'ملاحظات']
    },
    'positions': {
        'title': 'الوظائف والرواتب',
        'category': 'OPEX - الموارد البشرية',
        'columns': ['المسمى الوظيفي', 'الجنسية', 'العدد', 'الراتب الأساسي + السكن', 'أشهر العمل/سنة', 'إجمالي التكلفة (مع التأمينات)', 'متغير؟']
    },
    'logistics': {
        'title': 'الموارد اللوجستية',
        'category': 'OPEX - الموارد اللوجستية',
        'columns': ['البند', 'شهري', 'سنوي', '% متغير', 'ملاحظات']
    },
    'administrative': {
        'title': 'الموارد الإدارية',
        'category': 'OPEX - الموارد الإدارية',
        'columns': ['البند', 'شهري', 'سنوي', 'ملاحظات']
    },
    'licenses': {
        'title': 'التراخيص والرسوم القانونية',
        'category': 'CAPEX - الدراسة القانونية',
        'columns': ['البند', 'الكمية', 'السعر', 'الإجمالي', 'ملاحظات']
    },
    'competitors': {
        'title': 'تحليل المنافسين',
        'category': 'الدراسة التسويقية',
        'columns': ['المنافس', 'نقاط القوة', 'نقاط الضعف', 'الحصة السوقية %']
    },
    'campaigns': {
        'title': 'الحملات التسويقية',
        'category': 'الدراسة التسويقية',
        'columns': ['الحملة', 'النوع (رأسمالي/تشغيلي)', 'المبلغ', 'شهري (للتشغيلي)', 'ملاحظات']
    },
    'revenueStreams': {
        'title': 'مصادر الإيرادات',
        'category': 'الإيرادات',
        'columns': ['الخدمة', 'العملاء/شهر', 'متوسط السعر', 'نمو سنوي %', 'السنة 1']
    },
    'serviceItems': {
        'title': 'الخدمات المفصلة',
        'category': 'الإيرادات',
        'columns': ['اسم الخدمة', 'الرمز', 'CAPEX', 'تكاليف ثابتة/شهر', 'تكلفة/عميل', 'سعر الخدمة', 'عملاء/شهر', 'نمو سنوي %', 'الإيراد السنوي', 'الربح السنوي']
    },
    'pestelFactors': {
        'title': 'تحليل PESTEL',
        'category': 'التحليل الاستراتيجي',
        'columns': ['العامل', 'الوصف', 'التأثير', 'ملاحظات']
    },
    'swotItems': {
        'title': 'التحليل الرباعي (SWOT)',
        'category': 'التحليل الاستراتيجي',
        'columns': ['الفئة (قوة/ضعف/فرصة/تهديد)', 'البند', 'الأولوية']
    },
    'marketSegments': {
        'title': 'شرائح العملاء',
        'category': 'التحليل الاستراتيجي',
        'columns': ['الشريحة', 'الديموغرافيا', 'الاحتياجات', 'الحجم (ريال)', 'الأولوية']
    },
    'riskRegister': {
        'title': 'سجل المخاطر',
        'category': 'تحليل المخاطر',
        'columns': ['الخطر', 'النوع', 'الاحتمالية', 'الأثر', 'الدرجة (محسوبة)', 'خطة المواجهة', 'المسؤول']
    }
}

def create_excel_file():
    """إنشاء ملف Excel مع جميع الجداول"""
    
    # إنشاء workbook جديد
    wb = openpyxl.Workbook()
    
    # حذف الورقة الافتراضية
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ألوان للتصميم
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    category_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    category_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # تجميع الجداول حسب الفئة
    categories = {}
    for table_id, table_info in TABLES.items():
        category = table_info['category']
        if category not in categories:
            categories[category] = []
        categories[category].append((table_id, table_info))
    
    # ترتيب الفئات
    category_order = [
        'CAPEX - الدراسة الفنية',
        'CAPEX - الموارد التقنية',
        'CAPEX - الدراسة القانونية',
        'OPEX - الموارد البشرية',
        'OPEX - الموارد اللوجستية',
        'OPEX - الموارد الإدارية',
        'الدراسة التسويقية',
        'الإيرادات',
        'التحليل الاستراتيجي',
        'تحليل المخاطر'
    ]
    
    # إنشاء ورقة الفهرس
    index_sheet = wb.create_sheet('فهرس الجداول', 0)
    index_sheet['A1'] = 'فهرس جداول دراسة الجدوى'
    index_sheet['A1'].font = Font(bold=True, size=16)
    index_sheet['A1'].fill = category_fill
    index_sheet.merge_cells('A1:D1')
    
    row = 3
    index_sheet['A3'] = 'رقم'
    index_sheet['B3'] = 'اسم الجدول'
    index_sheet['C3'] = 'الفئة'
    index_sheet['D3'] = 'اسم الورقة'
    
    for cell in ['A3', 'B3', 'C3', 'D3']:
        index_sheet[cell].font = header_font
        index_sheet[cell].fill = header_fill
        index_sheet[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    row = 4
    sheet_num = 1
    
    # إنشاء ورقة لكل جدول
    for category in category_order:
        if category not in categories:
            continue
        
        for table_id, table_info in categories[category]:
            # إنشاء ورقة جديدة
            sheet_name = table_info['title'][:31]  # Excel limit: 31 chars
            ws = wb.create_sheet(sheet_name)
            
            # إضافة عنوان الجدول
            ws['A1'] = table_info['title']
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].fill = category_fill
            ws.merge_cells(f'A1:{get_column_letter(len(table_info["columns"]))}1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # إضافة الفئة
            ws['A2'] = f'الفئة: {category}'
            ws['A2'].font = category_font
            ws.merge_cells(f'A2:{get_column_letter(len(table_info["columns"]))}2')
            
            # إضافة رؤوس الأعمدة
            for col_idx, col_name in enumerate(table_info['columns'], start=1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = col_name
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
                
                # ضبط عرض العمود
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
            # إضافة صف فارغ للبيانات
            for col_idx in range(1, len(table_info['columns']) + 1):
                cell = ws.cell(row=4, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            
            # تجميد الصف الأول والثاني والثالث
            ws.freeze_panes = 'A4'
            
            # إضافة إلى الفهرس
            index_sheet[f'A{row}'] = sheet_num
            index_sheet[f'B{row}'] = table_info['title']
            index_sheet[f'C{row}'] = category
            index_sheet[f'D{row}'] = f'=HYPERLINK("#\'{sheet_name}\'!A1", "{sheet_name}")'
            
            for col in ['A', 'B', 'C', 'D']:
                cell = index_sheet[f'{col}{row}']
                cell.border = border
                cell.alignment = Alignment(horizontal='center' if col == 'A' else 'right', vertical='center')
            
            row += 1
            sheet_num += 1
    
    # تنسيق ورقة الفهرس
    index_sheet.column_dimensions['A'].width = 8
    index_sheet.column_dimensions['B'].width = 35
    index_sheet.column_dimensions['C'].width = 30
    index_sheet.column_dimensions['D'].width = 25
    
    # حفظ الملف
    import os
    filename = f'جداول_دراسة_الجدوى_{datetime.now().strftime("%Y%m%d")}.xlsx'
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    filepath = os.path.join(base_dir, filename)
    wb.save(filepath)
    
    print(f'✅ تم إنشاء الملف: {filepath}')
    print(f'📊 عدد الجداول: {len(TABLES)} جدول')
    print(f'📑 عدد الأوراق: {len(wb.sheetnames)} ورقة (بما فيها الفهرس)')
    
    return filepath

if __name__ == '__main__':
    try:
        import sys
        import os
        # الحصول على المسار الحالي للملف
        script_dir = os.path.dirname(os.path.abspath(__file__))
        parent_dir = os.path.dirname(script_dir)
        os.chdir(parent_dir)
        
        filepath = create_excel_file()
        print(f'\n✅ تم إنشاء الملف بنجاح!')
        print(f'📁 الموقع: {filepath}')
    except Exception as e:
        print(f'❌ خطأ: {e}')
        import traceback
        traceback.print_exc()
