import csv
import datetime as dt
import json
from collections import Counter

import openpyxl


def safe_unicode_escape(s: str) -> str:
    return s.encode("unicode_escape").decode("ascii")


def is_error_value(v) -> bool:
    return isinstance(v, str) and v.startswith("#")


def color_to_hex(col) -> str:
    if not col:
        return ""
    # Some openpyxl Color attributes can raise when accessed on malformed styles.
    try:
        if getattr(col, "type", None) == "rgb":
            rgb = getattr(col, "rgb", None)
            if rgb:
                return str(rgb)
    except Exception:
        pass
    try:
        indexed = getattr(col, "indexed", None)
        if isinstance(indexed, int):
            return f"indexed:{indexed}"
    except Exception:
        pass
    try:
        theme = getattr(col, "theme", None)
        if theme is not None:
            return f"theme:{theme}"
    except Exception:
        pass
    return ""


def cell_style_summary(c) -> dict:
    font = c.font
    fill = c.fill
    align = c.alignment
    border = c.border

    return {
        "number_format": c.number_format or "",
        "font_name": font.name or "",
        "font_size": font.sz if font.sz is not None else "",
        "font_bold": bool(font.b),
        "font_italic": bool(font.i),
        "font_underline": str(font.u or ""),
        "font_color": color_to_hex(font.color),
        "fill_type": fill.patternType or "",
        "fill_fg": color_to_hex(fill.fgColor),
        "fill_bg": color_to_hex(fill.bgColor),
        "align_h": align.horizontal or "",
        "align_v": align.vertical or "",
        "align_wrap": bool(align.wrap_text),
        "border_left": border.left.style or "",
        "border_right": border.right.style or "",
        "border_top": border.top.style or "",
        "border_bottom": border.bottom.style or "",
    }


def build_merged_map(ws) -> dict:
    merged_map = {}
    for r in ws.merged_cells.ranges:
        rng = str(r)
        for row in range(r.min_row, r.max_row + 1):
            for col in range(r.min_col, r.max_col + 1):
                merged_map[(row, col)] = rng
    return merged_map


def analyze(xlsx_path: str, report_path: str, dump_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    # workbook-level info
    named_ranges = []
    try:
        for dn in wb.defined_names.definedName:
            named_ranges.append(
                {
                    "name": dn.name,
                    "attr_text": dn.attr_text,
                    "localSheetId": dn.localSheetId,
                    "hidden": dn.hidden,
                }
            )
    except Exception:
        pass

    volatile_funcs = (
        "NOW(",
        "TODAY(",
        "RAND(",
        "RANDBETWEEN(",
        "OFFSET(",
        "INDIRECT(",
        "CELL(",
        "INFO(",
    )

    sheet_summaries = []
    formula_stats = Counter()
    error_cells = []
    external_ref_cells = []
    volatile_cells = []

    # Create cell dump
    with open(dump_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "sheet",
                "sheet_unicode_escape",
                "cell",
                "row",
                "col",
                "value",
                "data_type",
                "formula",
                "is_error",
                "hyperlink",
                "comment",
                "is_merged",
                "merged_range",
                "number_format",
                "font_name",
                "font_size",
                "font_bold",
                "font_italic",
                "font_underline",
                "font_color",
                "fill_type",
                "fill_fg",
                "fill_bg",
                "align_h",
                "align_v",
                "align_wrap",
                "border_left",
                "border_right",
                "border_top",
                "border_bottom",
            ]
        )

        for sname in wb.sheetnames:
            ws = wb[sname]
            dim = ws.calculate_dimension()  # e.g. A1:K120
            merged_map = build_merged_map(ws)

            # Parse dimension bounds
            min_col = ws.min_column
            min_row = ws.min_row
            max_col = ws.max_column
            max_row = ws.max_row

            # Some workbooks have inflated max_row/max_col; clamp using dimension when possible
            try:
                from openpyxl.utils.cell import range_boundaries

                a, b, c, d = range_boundaries(dim)
                min_col, min_row, max_col, max_row = a, b, c, d
            except Exception:
                pass

            nonempty_or_styled = 0
            formulas = 0
            errors = 0

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    val = cell.value
                    dtp = cell.data_type
                    formula = val if isinstance(val, str) and val.startswith("=") else ""

                    has_any = (
                        val not in (None, "")
                        or formula
                        or cell.has_style
                        or cell.hyperlink
                        or cell.comment
                    )
                    if has_any:
                        nonempty_or_styled += 1

                    if formula:
                        formulas += 1
                        formula_stats["total_formulas"] += 1
                        up = formula.upper()
                        if "[" in up and "]" in up:
                            external_ref_cells.append((sname, cell.coordinate, formula))
                            formula_stats["external_refs"] += 1
                        if any(fn in up for fn in volatile_funcs):
                            volatile_cells.append((sname, cell.coordinate, formula))
                            formula_stats["volatile"] += 1

                    if is_error_value(val):
                        errors += 1
                        error_cells.append((sname, cell.coordinate, val))

                    style = cell_style_summary(cell)

                    mrng = merged_map.get((row, col), "")
                    is_merged = bool(mrng)

                    w.writerow(
                        [
                            sname,
                            safe_unicode_escape(sname),
                            cell.coordinate,
                            row,
                            col,
                            "" if val is None else val,
                            dtp,
                            formula,
                            bool(is_error_value(val)),
                            (cell.hyperlink.target if cell.hyperlink else ""),
                            (cell.comment.text if cell.comment else ""),
                            is_merged,
                            mrng,
                            style["number_format"],
                            style["font_name"],
                            style["font_size"],
                            style["font_bold"],
                            style["font_italic"],
                            style["font_underline"],
                            style["font_color"],
                            style["fill_type"],
                            style["fill_fg"],
                            style["fill_bg"],
                            style["align_h"],
                            style["align_v"],
                            style["align_wrap"],
                            style["border_left"],
                            style["border_right"],
                            style["border_top"],
                            style["border_bottom"],
                        ]
                    )

            # sheet-level extras
            hidden_rows = []
            hidden_cols = []
            try:
                for r, dimobj in ws.row_dimensions.items():
                    if dimobj.hidden:
                        hidden_rows.append(r)
                for c, dimobj in ws.column_dimensions.items():
                    if dimobj.hidden:
                        hidden_cols.append(c)
            except Exception:
                pass

            # conditional formatting, data validation, auto filter
            cf_count = 0
            dv_count = 0
            try:
                cf_count = len(getattr(ws.conditional_formatting, "cf_rules", {}))
            except Exception:
                pass
            try:
                dv_count = (
                    len(ws.data_validations.dataValidation)
                    if getattr(ws, "data_validations", None)
                    else 0
                )
            except Exception:
                pass

            sheet_summaries.append(
                {
                    "name": sname,
                    "name_unicode_escape": safe_unicode_escape(sname),
                    "dimension": dim,
                    "min_row": min_row,
                    "min_col": min_col,
                    "max_row": max_row,
                    "max_col": max_col,
                    "rows": max_row - min_row + 1,
                    "cols": max_col - min_col + 1,
                    "nonempty_or_styled_cells_in_dim": nonempty_or_styled,
                    "formula_cells_in_dim": formulas,
                    "error_cells_in_dim": errors,
                    "merged_ranges_count": len(ws.merged_cells.ranges),
                    "hidden_rows_count": len(hidden_rows),
                    "hidden_cols_count": len(hidden_cols),
                    "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else "",
                    "auto_filter": str(ws.auto_filter.ref)
                    if ws.auto_filter and ws.auto_filter.ref
                    else "",
                    "conditional_formatting_rules": cf_count,
                    "data_validation_rules": dv_count,
                    "page_setup_orientation": str(ws.page_setup.orientation)
                    if ws.page_setup
                    else "",
                }
            )

    # Build markdown report
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = []
    lines.append("# تقرير تحليل ملف دراسة الجدوى (تلقائي)\\n\\n")
    lines.append(f"- وقت التحليل: **{now}**\\n")
    lines.append(f"- ملف المصدر: `{xlsx_path}`\\n")
    lines.append(f"- ملف تفريغ كل الخلايا (CSV): `{dump_path}`\\n")

    lines.append("\\n## ملخص عام\\n")
    lines.append(f"- عدد الشيتات: **{len(wb.sheetnames)}**\\n")
    lines.append(f"- عدد الـ Named Ranges: **{len(named_ranges)}**\\n")
    lines.append(f"- إجمالي الخلايا التي تحتوي صيغ: **{formula_stats.get('total_formulas', 0)}**\\n")
    lines.append(f"- خلايا تحتوي مراجع خارجية: **{formula_stats.get('external_refs', 0)}**\\n")
    lines.append(f"- خلايا تحتوي دوال متطايرة: **{formula_stats.get('volatile', 0)}**\\n")
    lines.append(f"- عدد خلايا أخطاء Excel (#DIV/0! وغيرها): **{len(error_cells)}**\\n")

    lines.append("\\n## الشيتات (الأسماء + نطاق الاستخدام)\\n")
    lines.append(
        "ملاحظة: تم حفظ اسم كل شيت أيضاً بصيغة `unicode_escape` لتفادي مشاكل الترميز في الطرفية.\\n\\n"
    )
    lines.append(
        "| # | Sheet | unicode_escape | Dimension | Rows x Cols | Merged | Freeze | Filter | CF | DV |\\n"
    )
    lines.append("|---:|---|---|---|---:|---:|---|---|---:|---:|\\n")
    for i, s in enumerate(sheet_summaries, 1):
        lines.append(
            f"| {i} | {s['name']} | `{s['name_unicode_escape']}` | `{s['dimension']}` | "
            f"{s['rows']} x {s['cols']} | {s['merged_ranges_count']} | {s['freeze_panes'] or '-'} | "
            f"{s['auto_filter'] or '-'} | {s['conditional_formatting_rules']} | {s['data_validation_rules']} |\\n"
        )

    lines.append("\\n## تدقيق الصيغ (Formulas)\\n")
    if external_ref_cells:
        lines.append("\\n### مراجع خارجية (قد تسبب مشاكل عند نقل الملف)\\n")
        lines.append("| Sheet (escape) | Cell | Formula |\\n|---|---|---|\\n")
        for sname, coord, formula in external_ref_cells[:200]:
            lines.append(f"| `{safe_unicode_escape(sname)}` | `{coord}` | `{formula}` |\\n")
        if len(external_ref_cells) > 200:
            lines.append(f"\\nتم عرض أول 200 فقط من أصل {len(external_ref_cells)}.\\n")
    else:
        lines.append("- لا توجد مراجع خارجية واضحة بصيغة `[Book.xlsx]Sheet!A1` داخل الصيغ.\\n")

    if volatile_cells:
        lines.append("\\n### دوال متطايرة (Volatile)\\n")
        lines.append("| Sheet (escape) | Cell | Formula |\\n|---|---|---|\\n")
        for sname, coord, formula in volatile_cells[:200]:
            lines.append(f"| `{safe_unicode_escape(sname)}` | `{coord}` | `{formula}` |\\n")
        if len(volatile_cells) > 200:
            lines.append(f"\\nتم عرض أول 200 فقط من أصل {len(volatile_cells)}.\\n")
    else:
        lines.append("\\n- لا توجد دوال متطايرة ضمن قائمة الفحص (NOW/TODAY/RAND/OFFSET/INDIRECT...).\\n")

    lines.append("\\n## خلايا الأخطاء\\n")
    if error_cells:
        lines.append("| Sheet (escape) | Cell | Value |\\n|---|---|---|\\n")
        for sname, coord, val in error_cells[:300]:
            lines.append(f"| `{safe_unicode_escape(sname)}` | `{coord}` | `{val}` |\\n")
        if len(error_cells) > 300:
            lines.append(f"\\nتم عرض أول 300 فقط من أصل {len(error_cells)}.\\n")
    else:
        lines.append("- لا توجد قيم أخطاء ظاهرة في الخلايا.\\n")

    lines.append("\\n## Named Ranges\\n")
    if named_ranges:
        lines.append("| Name | SheetId | Hidden | Ref |\\n|---|---:|---:|---|\\n")
        for nr in named_ranges[:300]:
            lines.append(
                f"| `{nr.get('name', '')}` | {nr.get('localSheetId', '')} | "
                f"{1 if nr.get('hidden') else 0} | `{nr.get('attr_text', '')}` |\\n"
            )
        if len(named_ranges) > 300:
            lines.append(f"\\nتم عرض أول 300 فقط من أصل {len(named_ranges)}.\\n")
    else:
        lines.append("- لا توجد Named Ranges (أو لا يمكن قراءتها عبر المكتبة).\\n")

    lines.append("\\n## ملاحظات مهمة (حدود التحليل)\\n")
    lines.append("- Excel قد يحتوي صيغ محسوبة؛ هذا التحليل يقرأ **الصيغة نفسها** ولا يعيد حساب النتائج.\\n")
    lines.append(
        "- إذا كان الملف يعتمد على ماكرو/PowerQuery/Pivot متقدم، قد لا يظهر كامل تفاصيله في هذا التقرير.\\n"
    )
    lines.append("- تم تفريغ **كل الخلايا داخل نطاق الاستخدام (Dimension)** لكل شيت داخل ملف CSV.\\n")

    with open(report_path, "w", encoding="utf-8") as rf:
        # NOTE: historical versions wrote literal "\\n" into the report.
        # Normalize to real newlines for correct Markdown rendering.
        rf.write("".join(lines).replace("\\n", "\n"))

    return {
        "xlsx_path": xlsx_path,
        "report_path": report_path,
        "dump_path": dump_path,
        "sheet_count": len(sheet_summaries),
        "sheets_unicode_escape": [s["name_unicode_escape"] for s in sheet_summaries],
        "total_formulas": int(formula_stats.get("total_formulas", 0)),
        "external_refs": int(formula_stats.get("external_refs", 0)),
        "volatile": int(formula_stats.get("volatile", 0)),
        "error_cells": len(error_cells),
    }


if __name__ == "__main__":
    # Keep all paths ASCII-safe by constructing them with unicode escapes in the runner.
    raise SystemExit(
        "Please run this script via the runner command that provides paths (see assistant)."
    )

