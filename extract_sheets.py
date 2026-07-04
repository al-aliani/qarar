import openpyxl
import json

wb = openpyxl.load_workbook(r'دراسة الجدوى لماك بلاش.xlsx', data_only=True)

# Key sheets to analyze
key_sheets = {
    "الدراسة المالية": 9,
    "الدراسة الفنية": 10,
    "تقدير الإيرادات": 11,
    "الدراسة القانونية": 13,
    "دراسة الموارد البشرية": 14,
    "دراسة الموارد التقنية": 15,
}

result = {}

for name, idx in key_sheets.items():
    ws = wb.worksheets[idx]
    print(f"\n{'='*50}")
    print(f"Sheet: {ws.title}")
    print(f"{'='*50}")
    
    # Extract first 15 rows
    rows_data = []
    for row in range(1, min(20, ws.max_row + 1)):
        row_vals = []
        for col in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row=row, column=col).value
            if cell is not None:
                row_vals.append(str(cell).strip()[:40])
            else:
                row_vals.append("")
        if any(row_vals):
            rows_data.append(row_vals)
            print(f"Row {row}: {row_vals}")
    
    result[name] = rows_data

with open('tables_detail.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
